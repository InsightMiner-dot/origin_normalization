import io
import json
import logging
import os
import time
from datetime import datetime
from typing import Dict, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from rapidfuzz import fuzz

DEFAULT_OUTPUT_SUFFIX = "_with_ct"
DEFAULT_LOG_FILE_NAME = "keywords_tagging_app.log"
DEFAULT_SHEET_NAME = "Consolidated data"

APP_DIR = os.path.dirname(os.path.abspath(__file__))
AUDIT_DIR = os.path.join(APP_DIR, "audit")
os.makedirs(AUDIT_DIR, exist_ok=True)
LOG_FILE_PATH = os.path.join(AUDIT_DIR, DEFAULT_LOG_FILE_NAME)

logging.basicConfig(
    filename=LOG_FILE_PATH,
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


def normalize_rules_payload(rules_data: Dict) -> Tuple[Dict[str, list], int]:
    if "rules" in rules_data:
        rules = rules_data["rules"]
        threshold = rules_data.get("config", {}).get("fuzzy_threshold", 85)
    else:
        rules = rules_data
        threshold = 85
    return rules, int(threshold)


def load_rules_from_upload(uploaded_rules_file) -> Tuple[Dict[str, list], int]:
    uploaded_rules_file.seek(0)
    rules_data = json.load(uploaded_rules_file)
    logger.info("Rules loaded from uploaded JSON file: %s", uploaded_rules_file.name)
    return normalize_rules_payload(rules_data)


def load_rules_from_path(rules_path: str) -> Tuple[Dict[str, list], int]:
    with open(rules_path, "r", encoding="utf-8") as handle:
        rules_data = json.load(handle)
    logger.info("Rules loaded from local path: %s", rules_path)
    return normalize_rules_payload(rules_data)


def build_keyword_map(rules: Dict[str, list]) -> Dict[str, str]:
    keyword_map = {}
    for charge_type, keyword_list in rules.items():
        for keyword in keyword_list:
            keyword_map[str(keyword).lower()] = charge_type
    return keyword_map


def classify_charge(description, uom, cost_rate, keyword_map: Dict[str, str], threshold: int) -> str:
    text = " ".join(
        [
            str(description).lower(),
            str(uom).lower(),
            str(cost_rate).lower(),
        ]
    )

    best_match = None
    best_score = 0

    for keyword, charge_type in keyword_map.items():
        score = fuzz.partial_ratio(keyword, text)
        if score > best_score:
            best_score = score
            best_match = charge_type

    return best_match if best_score >= threshold else "NEED_REVIEW"


def get_excel_sheet_names(uploaded_file) -> list[str]:
    uploaded_file.seek(0)
    excel_file = pd.ExcelFile(uploaded_file, engine="openpyxl")
    sheet_names = excel_file.sheet_names
    uploaded_file.seek(0)
    return sheet_names


def load_input_file(uploaded_file, sheet_name: str | None) -> pd.DataFrame:
    file_name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if file_name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, engine="openpyxl")


def build_output_filename(uploaded_filename: str) -> str:
    base_name, _ = os.path.splitext(os.path.basename(uploaded_filename))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}{DEFAULT_OUTPUT_SUFFIX}_{timestamp}.xlsx"


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return output.getvalue()


def workbook_to_excel_bytes(
    uploaded_file,
    result_df: pd.DataFrame,
    sheet_name: str,
    output_column: str,
) -> bytes:
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file)
    worksheet = workbook[sheet_name]

    header_row = 1
    header_to_col = {}
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value is not None:
            header_to_col[str(cell_value)] = col_idx

    if output_column in header_to_col:
        output_col_idx = header_to_col[output_column]
    else:
        output_col_idx = worksheet.max_column + 1
        worksheet.cell(row=header_row, column=output_col_idx).value = output_column

    for row_idx, value in enumerate(result_df[output_column].tolist(), start=2):
        worksheet.cell(row=row_idx, column=output_col_idx).value = value

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def process_dataframe(
    df: pd.DataFrame,
    keyword_map: Dict[str, str],
    threshold: int,
    description_column: str,
    uom_column: str,
    rate_column: str,
    output_column: str,
) -> pd.DataFrame:
    required_columns = [description_column, uom_column, rate_column]
    missing_columns = [column for column in required_columns if column not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")

    logger.info(
        "Keyword tagging started. Rows=%s, threshold=%s, output_column=%s",
        len(df),
        threshold,
        output_column,
    )

    result_df = df.copy()
    total = len(result_df)
    classifications = []

    start_time = time.time()
    progress_bar = st.progress(0, text="Preparing file processing...")
    status_text = st.empty()

    for idx, (_, row) in enumerate(result_df.iterrows(), start=1):
        charge_type = classify_charge(
            row.get(description_column, ""),
            row.get(uom_column, ""),
            row.get(rate_column, ""),
            keyword_map=keyword_map,
            threshold=threshold,
        )
        classifications.append(charge_type)

        elapsed_time = time.time() - start_time
        speed = (idx / elapsed_time) if elapsed_time > 0 else 0.0
        remaining_time = ((total - idx) / speed) if speed > 0 else 0.0

        progress_bar.progress(
            idx / total if total else 1.0,
            text="Processing file...",
        )
        status_text.text(
            f"Processed {idx}/{total} | Speed: {speed:.1f}/s | ETA: {remaining_time:.1f}s"
        )

    result_df[output_column] = classifications
    logger.info("Keyword tagging completed. Rows=%s", total)
    return result_df


def render_rules_summary(rules: Dict[str, list], threshold: int) -> None:
    total_keywords = sum(len(keywords) for keywords in rules.values())
    st.write(f"Charge types: {len(rules)}")
    st.write(f"Total keywords: {total_keywords}")
    st.write(f"Fuzzy threshold: {threshold}")


def main() -> None:
    st.set_page_config(page_title="Keywords Tagging Tool", layout="wide")
    st.title("Keywords Tagging Tool")
    st.caption("Upload a CSV or Excel file, load tagging rules, classify charge types, and download the enriched output.")

    with st.sidebar:
        st.header("Settings")
        description_column = st.text_input("Description column", value="Description")
        uom_column = st.text_input("UOM column", value="UOM (Volume)")
        rate_column = st.text_input("Rate column", value="Cost (Haul)/Rate")
        output_column = st.text_input("Output column", value="Charge_Type")
        st.caption(f"Audit log folder: `{AUDIT_DIR}`")

    uploaded_rules_file = st.file_uploader("Upload rules JSON", type=["json"])
    uploaded_file = st.file_uploader("Upload Excel or CSV", type=["xlsx", "xls", "csv"])

    if uploaded_rules_file is None:
        st.info("Upload a rules JSON file to continue.")
        return

    try:
        rules, threshold = load_rules_from_upload(uploaded_rules_file)
    except Exception as exc:
        logger.exception("Failed to load uploaded rules file.")
        st.error(f"Unable to read uploaded rules JSON: {exc}")
        return

    keyword_map = build_keyword_map(rules)

    with st.expander("Rules Summary", expanded=False):
        render_rules_summary(rules, threshold)

    if not uploaded_file:
        st.info("Upload a file to begin.")
        return

    logger.info("User uploaded tagging file: %s", uploaded_file.name)

    selected_sheet_name = None
    if uploaded_file.name.lower().endswith((".xlsx", ".xls")):
        try:
            sheet_names = get_excel_sheet_names(uploaded_file)
        except Exception as exc:
            logger.exception("Failed to inspect Excel sheets for uploaded file: %s", uploaded_file.name)
            st.error(f"Unable to inspect Excel sheets: {exc}")
            return

        if len(sheet_names) == 1:
            selected_sheet_name = sheet_names[0]
            st.info(f"Detected sheet: {selected_sheet_name}")
        else:
            default_index = sheet_names.index(DEFAULT_SHEET_NAME) if DEFAULT_SHEET_NAME in sheet_names else 0
            selected_sheet_name = st.selectbox("Select sheet", options=sheet_names, index=default_index)

    try:
        source_df = load_input_file(uploaded_file, selected_sheet_name)
    except Exception as exc:
        logger.exception("Failed to load uploaded tagging file: %s", uploaded_file.name)
        st.error(f"Unable to read the uploaded file: {exc}")
        return

    output_filename = build_output_filename(uploaded_file.name)

    st.success("File loaded successfully.")
    st.write(f"Rows: {len(source_df)}")
    st.write(f"Columns: {', '.join(source_df.columns.astype(str))}")
    st.write(f"Sheet: {selected_sheet_name or 'CSV'}")

    with st.expander("Preview input data", expanded=True):
        st.dataframe(source_df.head(20), use_container_width=True)

    if st.button("Run tagging", type="primary"):
        try:
            logger.info("Run tagging button clicked for file: %s", uploaded_file.name)
            with st.spinner("Classifying charge types..."):
                result_df = process_dataframe(
                    source_df,
                    keyword_map=keyword_map,
                    threshold=threshold,
                    description_column=description_column,
                    uom_column=uom_column,
                    rate_column=rate_column,
                    output_column=output_column,
                )
        except Exception as exc:
            logger.exception("Tagging failed for file: %s", uploaded_file.name)
            st.error(f"Processing failed: {exc}")
            return

        st.success("Tagging completed.")
        st.caption(f"Audit log file: {LOG_FILE_PATH}")
        st.caption(f"Download file name: {output_filename}")

        with st.expander("Preview output data", expanded=True):
            st.dataframe(result_df.head(50), use_container_width=True)

        if uploaded_file.name.lower().endswith(".xlsx") and selected_sheet_name:
            output_bytes = workbook_to_excel_bytes(
                uploaded_file,
                result_df=result_df,
                sheet_name=selected_sheet_name,
                output_column=output_column,
            )
        else:
            output_bytes = dataframe_to_excel_bytes(
                result_df,
                sheet_name=selected_sheet_name or "Tagged Data",
            )

        st.download_button(
            label="Download output Excel",
            data=output_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
